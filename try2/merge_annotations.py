# -*- coding: utf-8 -*-
"""
将 try2 下各书标注 xlsx/xls 转为 CSV，匹配章节、统一 tag，并合并为 merged.csv。
输出列：book, chapter, paragragh, tag（paragragh 为用户约定拼写）
导出前清洗：删除全角【】及其中的内容；删除「第…更」类更新标记。
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

TRY2 = Path(__file__).resolve().parent
ROOT = TRY2.parent

RAW = {
    "dazhuzai": ROOT / "raw_books" / "row_text" / "dazhuzai_all.txt",
    "yuanzun": ROOT / "raw_books" / "row_text" / "yuanzun_all.txt",
    "doupo": ROOT / "raw_books" / "doupo.txt",
    "wanxiang": ROOT / "raw_books" / "row_text" / "wanxiang_all.txt",
    "wudong": ROOT / "raw_books" / "row_text" / "wudong_all.txt",
}

CHAPTER_LINE_RE = re.compile(
    r"^第[一二三四五六七八九十百千万零〇两0-9]+章[^\n]*$", re.MULTILINE
)

_BRACKET_BLOCK = re.compile(r"【[^】]*】")
_GENG_MARK = re.compile(r"第[一二三四五六七八九十百千万零〇两0-9]+更")
_GENG_MARK_DIGIT = re.compile(r"第\d+更")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def split_by_chapters(text: str) -> list[tuple[str, int, int]]:
    """返回 [(章节标题行, 起始下标, 结束下标), ...]，按文中出现顺序。"""
    matches = list(CHAPTER_LINE_RE.finditer(text))
    if not matches:
        return []
    out: list[tuple[str, int, int]] = []
    for i, m in enumerate(matches):
        title = m.group(0).strip()
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        out.append((title, start, end))
    return out


def norm_for_match(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"\s+", " ", s)
    s = s.replace("\u3000", " ").strip()
    return s


def fingerprint(s: str, n: int = 100) -> str:
    s = norm_for_match(s)
    return s[:n] if len(s) >= n else s


def find_chapter_for_paragraph(
    full_text: str, chapters: list[tuple[str, int, int]], paragraph: str
) -> str | None:
    """在原文中定位段落所属章节标题行（多片段匹配，避免标注从句中截断导致首句对不上）。"""
    para = norm_for_match(paragraph)
    if len(para) < 12:
        return None

    chunks: list[str] = []
    for part in re.split(r"[。！？\n]+", para):
        t = part.strip()
        if len(t) >= 16:
            chunks.append(t[:120])
    # 滑动窗口：避免整句因个别错字无法命中，而句首子串仍可与原文对齐
    p2 = norm_for_match(paragraph)
    for i in range(0, min(len(p2), 400), 20):
        piece = p2[i : i + 40]
        if len(piece) >= 16:
            chunks.append(piece)
    if not chunks:
        chunks = [fingerprint(para, 100)]

    def try_chunks(body: str) -> bool:
        nb = norm_for_match(body)
        for ch in chunks:
            if len(ch) >= 12 and ch in nb:
                return True
        return False

    for title, a, b in chapters:
        body = full_text[a:b]
        if try_chunks(body):
            return title
    # 再试首段指纹
    fp = fingerprint(para, 80)
    if len(fp) >= 12:
        for title, a, b in chapters:
            body = norm_for_match(full_text[a:b])
            if fp in body:
                return title
    # 全文定位后再反查章节（处理标注与原文个别字词不一致导致分章匹配失败）
    for ch in chunks:
        if len(ch) < 16:
            continue
        pos = norm_for_match(full_text).find(ch)
        if pos == -1:
            pos = full_text.find(ch)
        if pos != -1:
            for title, a, b in chapters:
                if a <= pos < b:
                    return title
    return None


def build_yuanzun_chapter_map(full_text: str) -> dict[int, str]:
    """章节序号 -> 章节标题行（按出现顺序第 n 个章节标题）。"""
    chs = split_by_chapters(full_text)
    return {i + 1: t for i, (t, _, _) in enumerate(chs)}


def normalize_tag(raw) -> int:
    """统一为 1（扮猪/优越感/扮猪吃虎侧）或 2（占有感/占有侧）。"""
    if raw is None or (isinstance(raw, float) and raw != raw):
        return 1
    s = str(raw).strip()
    if s.isdigit():
        v = int(s)
        if v in (1, 2):
            return 1
        if v == 3:
            return 2
        return 1
    # 文本类标签（金手指归类为 2：占有/机缘获得感）
    if any(k in s for k in ("占有", "占有欲", "金手指")):
        return 2
    if any(
        k in s
        for k in (
            "扮猪",
            "优越感",
            "吃虎",
            "先抑",
            "扬",
            "畅快",
            "机缘",
        )
    ):
        return 1
    return 1


def clean_export_text(s: str) -> str:
    """删除【】及其中内容；删除「第〇更」类字样；整理空白。"""
    if not s:
        return ""
    t = str(s)
    while True:
        n = _BRACKET_BLOCK.sub("", t)
        if n == t:
            break
        t = n
    t = _GENG_MARK.sub("", t)
    t = _GENG_MARK_DIGIT.sub("", t)
    t = re.sub(r"[ \t\u3000]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


def strip_wanxiang_noise(text: str) -> str:
    """去掉常见爬取杂质。"""
    if not text:
        return text
    lines = []
    for line in text.split("\n"):
        t = line.strip()
        if re.search(r"小说中还有哪些类似的情节", t):
            break
        if re.search(r"请再帮我处理", t):
            break
        if re.search(r"如何快速去除小说", t):
            break
        lines.append(line)
    return "\n".join(lines).strip()


def write_csv(path: Path, rows: list[tuple[str, str, int]], book_name: str) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["book", "chapter", "paragragh", "tag"])
        for ch, para, tag in rows:
            w.writerow([book_name, ch, para, tag])


def process_dazhuzai() -> list[tuple[str, str, int]]:
    text = read_text(RAW["dazhuzai"])
    chapters = split_by_chapters(text)
    wb = load_workbook(TRY2 / "dazhuzai_ann.xlsx", read_only=True)
    ws = wb.active
    out: list[tuple[str, str, int]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        para, tag_raw = row[0], row[1]
        if para is None:
            continue
        p = str(para).strip()
        if not p:
            continue
        ch = find_chapter_for_paragraph(text, chapters, p)
        if ch is None:
            ch = "未匹配章节"
        tag = normalize_tag(tag_raw)
        out.append((clean_export_text(ch), clean_export_text(p), tag))
    return out


def process_yuanzun() -> list[tuple[str, str, int]]:
    text = read_text(RAW["yuanzun"])
    cmap = build_yuanzun_chapter_map(text)
    wb = load_workbook(TRY2 / "yuanzun_ann.xlsx", read_only=True)
    ws = wb.active
    out: list[tuple[str, str, int]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        ch_raw, para, tag_raw = row[0], row[1], row[2]
        if para is None:
            continue
        p = str(para).strip()
        if not p:
            continue
        try:
            n = int(float(ch_raw)) if ch_raw is not None else None
        except (TypeError, ValueError):
            n = None
        if n is not None and n in cmap:
            ch = cmap[n]
        elif n is not None:
            ch = f"第{n}章(原文未找到标题)"
        else:
            ch = str(ch_raw) if ch_raw is not None else ""
        tag = normalize_tag(tag_raw)
        out.append((ch, p, tag))
    return out


def process_doupo() -> list[tuple[str, str, int]]:
    text = read_text(RAW["doupo"])
    chapters = split_by_chapters(text)
    wb = load_workbook(TRY2 / "doupo_ann.xlsx", read_only=True)
    ws = wb.active
    out: list[tuple[str, str, int]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        ch_raw, para, tag_raw = row[0], row[1], row[2]
        if para is None:
            continue
        p = str(para).strip()
        if not p:
            continue
        ch = str(ch_raw).strip() if ch_raw is not None else ""
        # 章节列异常时尝试用原文匹配
        if not ch or ch in ("扮猪", "占有感") or not ch.startswith("第"):
            mch = find_chapter_for_paragraph(text, chapters, p)
            if mch:
                ch = mch
        tag = normalize_tag(tag_raw)
        out.append((ch, p, tag))
    return out


def process_wanxiang() -> list[tuple[str, str, int]]:
    wb = load_workbook(TRY2 / "wanxiang_ann.xlsx", read_only=True)
    ws = wb.active
    out: list[tuple[str, str, int]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        ch_raw, para, tag_raw = row[0], row[1], row[2]
        if para is None:
            continue
        p = strip_wanxiang_noise(str(para).strip())
        if not p:
            continue
        ch = str(ch_raw).strip() if ch_raw is not None else ""
        tag = normalize_tag(tag_raw)
        out.append((clean_export_text(ch), clean_export_text(p), tag))
    return out


def process_wudong_xls() -> list[tuple[str, str, int]]:
    import xlrd

    book = xlrd.open_workbook(str(TRY2 / "wdqk.xls"))
    sh = book.sheet_by_index(0)
    out: list[tuple[str, str, int]] = []
    for r in range(1, sh.nrows):
        ch = str(sh.cell_value(r, 0)).strip()
        para = str(sh.cell_value(r, 1)).strip()
        tag_raw = sh.cell_value(r, 2)
        if not para:
            continue
        tag = normalize_tag(tag_raw)
        out.append((ch, para, tag))
    return out


def main() -> None:
    # 各书处理并写 CSV
    bundles = [
        ("dazhuzai_ann.csv", process_dazhuzai, "大主宰"),
        ("yuanzun_ann.csv", process_yuanzun, "元尊"),
        ("doupo_ann.csv", process_doupo, "斗破"),
        ("wanxiang_ann.csv", process_wanxiang, "万相"),
        ("wdqk_ann.csv", process_wudong_xls, "武动"),
    ]

    merged: list[tuple[str, str, str, int]] = []

    for csv_name, fn, book_name in bundles:
        rows = fn()
        write_csv(TRY2 / csv_name, rows, book_name)
        merged.extend([(book_name, ch, para, tag) for ch, para, tag in rows])

    merged_path = TRY2 / "merged.csv"
    with merged_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["book", "chapter", "paragragh", "tag"])
        for book_name, ch, para, tag in merged:
            w.writerow([book_name, ch, para, tag])

    print(
        "完成：",
        "各书 CSV + merged.csv，",
        f"合并行数 {len(merged)}",
    )


if __name__ == "__main__":
    main()
